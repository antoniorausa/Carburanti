import io
import os
import csv
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.chart.label import DataLabel, DataLabelList

RIGA_INTESTAZIONE = 8

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MAPPING_CSV = os.path.join(BASE_DIR, "mapping.csv")
PDV_CSV     = os.path.join(BASE_DIR, "pdv_selezionati.csv")

COLORI_CATEGORIA = {
    "CAD":       "FF9999",
    "CCN":       "99FF99",
    "CIA":       "FFFF99",
    "CNO":       "99CCFF",
    "PAC 2000A": "FFC000",
}

COLORI_CARBURANTI = {
    "Gasolio": "ED7D31",
    "Benzina": "70AD47",
    "GPL":     "4472C4",
    "Metano":  "E9967A",
}

NOMI_CARBURANTI = ["Gasolio", "Benzina", "GPL", "Metano"]


def _pulisci_nome_foglio(nome):
    for ch in r"/\:*?[]'":
        nome = nome.replace(ch, "-")
    return nome[:31]


def _leggi_mapping_csv():
    mapping, colori, distanze, carburanti = {}, {}, {}, {}
    if not os.path.exists(MAPPING_CSV):
        return mapping, colori, distanze, carburanti
    with open(MAPPING_CSV, newline="", encoding="utf-8-sig") as f:
        sample = f.read(1024); f.seek(0)
        sep = "\t" if "\t" in sample else ";"
        reader = csv.reader(f, delimiter=sep)
        next(reader, None)
        for row in reader:
            if not row or not row[0].strip():
                continue
            k = row[0].strip()
            mapping[k]  = row[1].strip() if len(row) > 1 else ""
            colori[k]   = row[2].strip().upper() if len(row) > 2 else ""
            distanze[k] = row[3].strip() if len(row) > 3 else ""
            carburanti[k] = [
                int(row[4]) if len(row) > 4 and row[4].strip() else 1,
                int(row[5]) if len(row) > 5 and row[5].strip() else 1,
                int(row[6]) if len(row) > 6 and row[6].strip() else 1,
                int(row[7]) if len(row) > 7 and row[7].strip() else 1,
            ]
    return mapping, colori, distanze, carburanti


def _leggi_pdv_csv():
    pdv = set()
    if not os.path.exists(PDV_CSV):
        return pdv
    with open(PDV_CSV, newline="", encoding="utf-8-sig") as f:
        sample = f.read(1024); f.seek(0)
        sep = "\t" if "\t" in sample else ";"
        reader = csv.reader(f, delimiter=sep)
        next(reader, None)
        for row in reader:
            if row and row[0].strip():
                pdv.add(row[0].strip())
    return pdv


def _parse_distanza(val):
    if val is None:
        return 0.0
    s = str(val).replace("'", "").replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_prezzo(val):
    """
    Gestisce sia prezzi già decimali (1.709) sia interi gonfiati (1709).
    Se il valore è > 10, lo divide per 1000 (es. 1709 → 1.709).
    """
    if val is None or str(val).strip() in ("-", "", "None"):
        return 0.0
    s = str(val).replace("'", "").replace(",", ".").strip()
    try:
        v = float(s)
        if v > 10:
            v = v / 1000.0
        return round(v, 3)
    except ValueError:
        return 0.0


def _leggi_sorgente_xls(file_bytes):
    sniff = file_bytes[:20].strip().lower()
    if sniff.startswith(b"<") or b"<!do" in sniff or b"<html" in sniff:
        return _leggi_sorgente_html_xls(file_bytes)
    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheet_by_index(0)
    skip = RIGA_INTESTAZIONE - 1
    rows = []
    for i in range(skip, ws.nrows):
        row = []
        for j in range(ws.ncols):
            cell = ws.cell(i, j)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                v = cell.value
                row.append(int(v) if v == int(v) else v)
            else:
                row.append(str(cell.value).strip() if cell.value != "" else None)
        rows.append(row)
    return rows


def _leggi_sorgente_html_xls(file_bytes):
    import pandas as pd
    skip = RIGA_INTESTAZIONE - 1
    try:
        tables = pd.read_html(io.BytesIO(file_bytes), header=0, skiprows=skip - 1, encoding="utf-8")
    except Exception:
        tables = pd.read_html(io.BytesIO(file_bytes), header=0, skiprows=skip - 1, encoding="latin-1")
    df = tables[0]
    rows = [list(df.columns)]
    for _, r in df.iterrows():
        rows.append([None if (str(v) in ("nan", "None", "")) else v for v in r])
    return rows


def _leggi_sorgente_xlsx(file_bytes):
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]
    all_rows = list(ws.values)
    skip = RIGA_INTESTAZIONE - 1
    return [list(r) for r in all_rows[skip:]]


def _hex_to_rgb_tuple(hex_color):
    h = hex_color.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))


def processa_excel(file_bytes, filename="file.xls"):
    mapping, colori, distanze, carburanti = _leggi_mapping_csv()
    pdv_selezionati = _leggi_pdv_csv()

    is_xls = filename.lower().endswith(".xls") and not filename.lower().endswith(".xlsx")
    src_rows = _leggi_sorgente_xls(file_bytes) if is_xls else _leggi_sorgente_xlsx(file_bytes)
    data_rows = src_rows[1:]

    IDX_CODICE   = 0
    IDX_COMUNE   = 1
    IDX_DISTANZA = 6
    IDX_PREZZI   = [7, 8, 9, 10]

    coppie_viste = {}
    for r in data_rows:
        k1 = str(r[IDX_CODICE]).strip() if r[IDX_CODICE] else ""
        k2 = str(r[IDX_COMUNE]).strip() if r[IDX_COMUNE] else ""
        chiave = f"{k1}|{k2}"
        if k1 and k2 and chiave not in coppie_viste:
            coppie_viste[chiave] = (k1, k2)

    coppie_ordinate = sorted(
        coppie_viste.values(),
        key=lambda x: mapping.get(x[0], "ZZ_" + x[0])
    )

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    ws_indice = wb_out.create_sheet("Indice", 0)
    ws_indice["A1"] = "Indice Fogli Generati"
    ws_indice["A1"].font = Font(bold=True)
    riga_indice = 2
    nomi_fogli_creati = []
    titoli_grafici = {}  # chart_filename -> titolo
    chart_counter = 1

    for (v1, v2) in coppie_ordinate:
        nome_da_mapping = mapping.get(v1, f"{v1} - {v2}")
        nome_foglio = _pulisci_nome_foglio(nome_da_mapping)
        base = nome_foglio
        suf = 1
        while nome_foglio in nomi_fogli_creati:
            nome_foglio = f"{base[:28]}_{suf}"
            suf += 1
        nomi_fogli_creati.append(nome_foglio)

        righe_filtrate = [
            r for r in data_rows
            if str(r[IDX_CODICE]).strip() == v1 and str(r[IDX_COMUNE]).strip() == v2
        ]

        flags = carburanti.get(v1, [1, 1, 1, 1])
        carb_attivi = [(i, NOMI_CARBURANTI[i]) for i, f in enumerate(flags) if f == 1]

        soglia = distanze.get(v1, "")
        try:
            soglia_f = float(str(soglia).replace(",", ".")) if soglia != "" else None
        except (ValueError, TypeError):
            soglia_f = None

        righe_elaborate = []
        for r in righe_filtrate:
            dist = _parse_distanza(r[IDX_DISTANZA])
            prezzi = [_parse_prezzo(r[IDX_PREZZI[i]]) for i in range(4)]
            somma_attivi = sum(prezzi[i] for i, f in enumerate(flags) if f == 1)
            if somma_attivi == 0:
                continue
            if soglia_f is not None and dist > soglia_f:
                continue
            riga_out = [
                str(r[3]).strip() if r[3] else "",
                str(r[4]).strip() if r[4] else "",
                str(r[5]).strip() if r[5] else "",
                dist,
            ]
            for i, _ in carb_attivi:
                riga_out.append(prezzi[i] if prezzi[i] > 0 else None)
            righe_elaborate.append(riga_out)

        righe_elaborate.sort(key=lambda x: x[3])

        # ── Scrivi foglio ────────────────────────────────────────────────────
        ws = wb_out.create_sheet(nome_foglio)

        cat = colori.get(v1, "")
        colore_hex = COLORI_CATEGORIA.get(cat, None)
        if colore_hex:
            ws.sheet_properties.tabColor = colore_hex

        cell_idx = ws_indice.cell(row=riga_indice, column=1, value=nome_foglio)
        cell_idx.hyperlink = f"#{nome_foglio}!A1"
        cell_idx.font = Font(color="0000FF", underline="single")
        if colore_hex:
            cell_idx.fill = PatternFill("solid", fgColor=colore_hex)
        riga_indice += 1

        intestazione = ["Insegna", "Comune", "Indirizzo", "Distanza (min)"] + \
                       [n for _, n in carb_attivi]
        for ci, h in enumerate(intestazione, 1):
            ws.cell(row=1, column=ci, value=h).font = Font(bold=True)

        righe_gialle = []
        for ri, riga in enumerate(righe_elaborate, 2):
            is_pdv = riga[2] in pdv_selezionati
            for ci, val in enumerate(riga, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if ci >= 5:
                    cell.number_format = '0.000'
                if is_pdv:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
            if is_pdv:
                righe_gialle.append(ri)

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

        # ── Grafico ──────────────────────────────────────────────────────────
        ultima_riga_dati = len(righe_elaborate) + 1
        if ultima_riga_dati > 1 and carb_attivi:
            righe_grafico = righe_gialle if righe_gialle else list(range(2, ultima_riga_dati + 1))
            n_punti = len(righe_grafico) + 1  # +1 per MEDIA

            nome_ws_g = f"_g_{nome_foglio}"[:31]
            ws_g = wb_out.create_sheet(nome_ws_g)
            ws_g.sheet_state = "hidden"

            # Riga 1: etichette (Insegna - Indirizzo, Comune)
            ws_g.cell(1, 1, "Serie")
            for ci, ri in enumerate(righe_grafico, 2):
                insegna   = ws.cell(ri, 1).value or ""
                indirizzo = ws.cell(ri, 3).value or ""
                comune_conc = ws.cell(ri, 2).value or ""
                # Formato: Insegna - Indirizzo comp., Comune (comune del distributore v2)
                lbl = f"{insegna} - {indirizzo}, {comune_conc}"
                ws_g.cell(1, ci, lbl)
            ws_g.cell(1, n_punti + 1, "MEDIA")

            # Righe 2+: valori per carburante + media
            for s_idx, (orig_idx, nome_carb) in enumerate(carb_attivi):
                col_src = 5 + s_idx
                riga_g = s_idx + 2
                ws_g.cell(riga_g, 1, nome_carb)
                valori = []
                for ci, ri in enumerate(righe_grafico, 2):
                    v = ws.cell(ri, col_src).value
                    v = round(float(v), 3) if v is not None and str(v) not in ("", "None") else 0.0
                    ws_g.cell(riga_g, ci, v)
                    valori.append((ri, v))

                # Media esclude distanza 0
                vals_media = [v for ri, v in valori if v > 0 and (ws.cell(ri, 4).value or 0) != 0]
                media = round(sum(vals_media) / len(vals_media), 3) if vals_media else 0.0
                ws_g.cell(riga_g, n_punti + 1, media)

            # ── Costruzione grafico ─────────────────────────────────────────
            from openpyxl.chart.title import Title
            from openpyxl.chart.text import RichText as ChartRichText
            from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties

            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.gapWidth = 100
            chart.width = 30
            chart.height = 18
            chart.shape = 4
            chart.style = 8  # tema scuro Excel stile 8

            # Titolo con overlay=False, testo nero visibile
            from openpyxl.chart.text import Text
            from openpyxl.drawing.text import RegularTextRun
            t_rpr = CharacterProperties(b=True, solidFill="000000")
            t_run = RegularTextRun(t=f"Confronto Prezzi Carburanti - [{nome_foglio}]", rPr=t_rpr)
            t_p   = Paragraph(pPr=ParagraphProperties(defRPr=t_rpr), r=[t_run])
            chart.title = Title(tx=Text(rich=ChartRichText(p=[t_p])), overlay=False)

            # Asse Y: parte da 0.5, formato 0,000
            from openpyxl.chart.axis import ChartLines
            chart.y_axis.numFmt = '0.000'
            chart.y_axis.scaling.min = 0.5
            chart.y_axis.majorGridlines = ChartLines()
            chart.y_axis.delete = False

            # Asse X: nascondi etichette tick (le mettiamo nel data table)
            chart.x_axis.tickLblPos = "none"
            chart.x_axis.delete = False

            # Data table in basso
            from openpyxl.chart.plotarea import DataTable
            chart.plot_area.dTable = DataTable(showHorzBorder=True, showVertBorder=True, showOutline=True, showKeys=True)

            # Categorie testuali embedded (StrRef) per mostrare le etichette nel grafico
            from openpyxl.chart.data_source import StrRef, StrVal, StrData, AxDataSource
            lbl_pts = []
            for ci in range(2, n_punti + 2):
                v = ws_g.cell(1, ci).value or ""
                lbl_pts.append(StrVal(idx=ci - 2, v=str(v)))
            ax_src = AxDataSource(strRef=StrRef(strCache=StrData(ptCount=len(lbl_pts), pt=lbl_pts)))

            for s_idx, (orig_idx, nome_carb) in enumerate(carb_attivi):
                riga_g = s_idx + 2
                colore_carb = COLORI_CARBURANTI.get(nome_carb, "4472C4")

                # Leggi valori per trovare min/max
                vals_riga = []
                for ci in range(2, n_punti + 2):
                    v = ws_g.cell(riga_g, ci).value or 0
                    vals_riga.append(float(v))

                # Esclude la colonna MEDIA (ultima) per trovare min/max sui dati reali
                vals_solo_dati = [v for v in vals_riga[:-1] if v > 0]
                val_max = max(vals_solo_dati) if vals_solo_dati else None
                val_min = min(vals_solo_dati) if vals_solo_dati else None

                data_ref = Reference(ws_g, min_col=2, max_col=n_punti + 1, min_row=riga_g, max_row=riga_g)
                ser = Series(data_ref, title=nome_carb)
                ser.graphicalProperties.solidFill = colore_carb
                ser.graphicalProperties.line.solidFill = colore_carb
                ser.cat = ax_src

                # Data labels: mostra solo min, max e MEDIA (ultima colonna)
                labels = []
                for pt_idx, v in enumerate(vals_riga):
                    if v == 0:
                        continue
                    is_max   = (v == val_max)
                    is_min   = (v == val_min)
                    is_media = (pt_idx == len(vals_riga) - 1)

                    if is_max or is_min or is_media:
                        if is_max:
                            txPr = _make_label_txpr("FFFFFF", bold=True)
                            spPr = _make_label_sppr("FF0000")
                        elif is_min:
                            txPr = _make_label_txpr("000000", bold=True)
                            spPr = _make_label_sppr("00FF00")
                        else:  # media
                            txPr = _make_label_txpr("FF0000", bold=True, italic=True)
                            spPr = _make_label_sppr("FFFF00")

                        dl = DataLabel(
                            idx=pt_idx,
                            showVal=True,
                            showLegendKey=False,
                            showCatName=False,
                            showSerName=False,
                            dLblPos="outEnd",
                            txPr=txPr,
                            spPr=spPr,
                        )
                        labels.append(dl)

                ser.dLbls = DataLabelList(
                    dLbl=labels,
                    showVal=False,
                    showLegendKey=False,
                    showCatName=False,
                    showSerName=False,
                )

                chart.series.append(ser)

            chart.legend = None  # Nessuna legenda (usa data table)

            col_chart = get_column_letter(len(intestazione) + 2)
            ws.add_chart(chart, f"{col_chart}1")
            titoli_grafici[f'chart{chart_counter}.xml'] = f'Confronto Prezzi Carburanti - [{nome_foglio}]'
            chart_counter += 1

    ws_indice.column_dimensions["A"].width = 35

    # Salva in buffer temporaneo
    tmp = io.BytesIO()
    wb_out.save(tmp)
    tmp.seek(0)

    # Post-processing: applica sfondo scuro a tutti i grafici via XML
    result = _applica_tema_scuro(tmp.read(), titoli_grafici)
    return result, len(nomi_fogli_creati)


# ── Costanti estratte dal template grafico Excel ─────────────────────────────
_TPL_ALTCONTENT    = '<mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"><mc:Choice Requires="c14" xmlns:c14="http://schemas.microsoft.com/office/drawing/2007/8/2/chart"><c14:style val="102"/></mc:Choice><mc:Fallback><c:style val="2"/></mc:Fallback></mc:AlternateContent>'
_TPL_CHARTSPACE_SPPR = '<c:spPr><a:gradFill flip="none" rotWithShape="1"><a:gsLst><a:gs pos="0"><a:schemeClr val="dk1"><a:lumMod val="65000"/><a:lumOff val="35000"/></a:schemeClr></a:gs><a:gs pos="100000"><a:schemeClr val="dk1"><a:lumMod val="85000"/><a:lumOff val="15000"/></a:schemeClr></a:gs></a:gsLst><a:path path="circle"><a:fillToRect l="50000" t="50000" r="50000" b="50000"/></a:path><a:tileRect/></a:gradFill><a:ln><a:noFill/></a:ln><a:effectLst/></c:spPr>'
_TPL_CATAX         = '<c:catAx><c:axId val="1147168719"/><c:scaling><c:orientation val="minMax"/></c:scaling><c:delete val="0"/><c:axPos val="b"/><c:numFmt formatCode="General" sourceLinked="1"/><c:majorTickMark val="none"/><c:minorTickMark val="none"/><c:tickLblPos val="nextTo"/><c:spPr><a:noFill/><a:ln w="12700" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="54000"/></a:schemeClr></a:solidFill><a:round/></a:ln><a:effectLst/></c:spPr><c:txPr><a:bodyPr rot="-60000000" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1"/><a:lstStyle/><a:p><a:pPr><a:defRPr sz="900" b="0" i="0" u="none" strike="noStrike" kern="1200" baseline="0"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="85000"/></a:schemeClr></a:solidFill><a:latin typeface="+mn-lt"/><a:ea typeface="+mn-ea"/><a:cs typeface="+mn-cs"/></a:defRPr></a:pPr><a:endParaRPr lang="it-IT"/></a:p></c:txPr><c:crossAx val="1147152399"/><c:crosses val="autoZero"/><c:auto val="1"/><c:lblAlgn val="ctr"/><c:lblOffset val="100"/><c:noMultiLvlLbl val="0"/></c:catAx>'
_TPL_VALAX         = '<c:valAx><c:axId val="1147152399"/><c:scaling><c:orientation val="minMax"/><c:min val="0.5"/></c:scaling><c:delete val="0"/><c:axPos val="l"/><c:majorGridlines><c:spPr><a:ln w="9525" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="10000"/></a:schemeClr></a:solidFill><a:round/></a:ln><a:effectLst/></c:spPr></c:majorGridlines><c:numFmt formatCode="0.000" sourceLinked="0"/><c:majorTickMark val="none"/><c:minorTickMark val="none"/><c:tickLblPos val="nextTo"/><c:spPr><a:noFill/><a:ln><a:noFill/></a:ln><a:effectLst/></c:spPr><c:txPr><a:bodyPr rot="-60000000" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1"/><a:lstStyle/><a:p><a:pPr><a:defRPr sz="900" b="0" i="0" u="none" strike="noStrike" kern="1200" baseline="0"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="85000"/></a:schemeClr></a:solidFill><a:latin typeface="+mn-lt"/><a:ea typeface="+mn-ea"/><a:cs typeface="+mn-cs"/></a:defRPr></a:pPr><a:endParaRPr lang="it-IT"/></a:p></c:txPr><c:crossAx val="1147168719"/><c:crosses val="autoZero"/><c:crossBetween val="between"/></c:valAx>'
_TPL_DTABLE        = '<c:dTable><c:showHorzBorder val="1"/><c:showVertBorder val="1"/><c:showOutline val="1"/><c:showKeys val="1"/><c:spPr><a:noFill/><a:ln w="9525"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="54000"/></a:schemeClr></a:solidFill></a:ln><a:effectLst/></c:spPr><c:txPr><a:bodyPr rot="0" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1"/><a:lstStyle/><a:p><a:pPr rtl="0"><a:defRPr sz="900" b="0" i="0" u="none" strike="noStrike" kern="1200" baseline="0"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="85000"/></a:schemeClr></a:solidFill><a:latin typeface="+mn-lt"/><a:ea typeface="+mn-ea"/><a:cs typeface="+mn-cs"/></a:defRPr></a:pPr><a:endParaRPr lang="it-IT"/></a:p></c:txPr></c:dTable>'
_TPL_TITLE         = '<c:title><c:tx><c:rich><a:bodyPr rot="0" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1"/><a:lstStyle/><a:p><a:pPr><a:defRPr sz="1600" b="1" i="0" u="none" strike="noStrike" kern="1200" spc="100" baseline="0"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/></a:schemeClr></a:solidFill><a:effectLst><a:outerShdw blurRad="50800" dist="38100" dir="5400000" algn="t" rotWithShape="0"><a:prstClr val="black"><a:alpha val="40000"/></a:prstClr></a:outerShdw></a:effectLst><a:latin typeface="+mn-lt"/><a:ea typeface="+mn-ea"/><a:cs typeface="+mn-cs"/></a:defRPr></a:pPr><a:r><a:rPr lang="it-IT"/><a:t>__TITLE__</a:t></a:r></a:p></c:rich></c:tx><c:overlay val="0"/><c:spPr><a:noFill/><a:ln><a:noFill/></a:ln><a:effectLst/></c:spPr><c:txPr><a:bodyPr rot="0" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1"/><a:lstStyle/><a:p><a:pPr><a:defRPr sz="1600" b="1" i="0" u="none" strike="noStrike" kern="1200" spc="100" baseline="0"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/></a:schemeClr></a:solidFill><a:effectLst><a:outerShdw blurRad="50800" dist="38100" dir="5400000" algn="t" rotWithShape="0"><a:prstClr val="black"><a:alpha val="40000"/></a:prstClr></a:outerShdw></a:effectLst><a:latin typeface="+mn-lt"/><a:ea typeface="+mn-ea"/><a:cs typeface="+mn-cs"/></a:defRPr></a:pPr><a:endParaRPr lang="it-IT"/></a:p></c:txPr></c:title>'
_TPL_STYLE1        = '<cs:chartStyle xmlns:cs="http://schemas.microsoft.com/office/drawing/2012/chartStyle" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" id="209"><cs:axisTitle><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"><a:lumMod val="85000"/></a:schemeClr></cs:fontRef><cs:defRPr sz="900" b="1" kern="1200" cap="all"/></cs:axisTitle><cs:categoryAxis><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"><a:lumMod val="85000"/></a:schemeClr></cs:fontRef><cs:spPr><a:ln w="12700" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="54000"/></a:schemeClr></a:solidFill><a:round/></a:ln></cs:spPr><cs:defRPr sz="900" kern="1200"/></cs:categoryAxis><cs:chartArea><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="dk1"/></cs:fontRef><cs:spPr><a:gradFill flip="none" rotWithShape="1"><a:gsLst><a:gs pos="0"><a:schemeClr val="dk1"><a:lumMod val="65000"/><a:lumOff val="35000"/></a:schemeClr></a:gs><a:gs pos="100000"><a:schemeClr val="dk1"><a:lumMod val="85000"/><a:lumOff val="15000"/></a:schemeClr></a:gs></a:gsLst><a:path path="circle"><a:fillToRect l="50000" t="50000" r="50000" b="50000"/></a:path><a:tileRect/></a:gradFill></cs:spPr><cs:defRPr sz="1000" kern="1200"/></cs:chartArea><cs:dataLabel><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"><a:lumMod val="85000"/></a:schemeClr></cs:fontRef><cs:defRPr sz="900" kern="1200"/></cs:dataLabel><cs:dataLabelCallout><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="dk1"><a:lumMod val="65000"/><a:lumOff val="35000"/></a:schemeClr></cs:fontRef><cs:spPr><a:solidFill><a:schemeClr val="lt1"/></a:solidFill></cs:spPr><cs:defRPr sz="900" kern="1200"/><cs:bodyPr rot="0" spcFirstLastPara="1" vertOverflow="clip" horzOverflow="clip" vert="horz" wrap="square" lIns="36576" tIns="18288" rIns="36576" bIns="18288" anchor="ctr" anchorCtr="1"><a:spAutoFit/></cs:bodyPr></cs:dataLabelCallout><cs:dataPoint><cs:lnRef idx="0"/><cs:fillRef idx="3"><cs:styleClr val="auto"/></cs:fillRef><cs:effectRef idx="3"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef></cs:dataPoint><cs:dataPoint3D><cs:lnRef idx="0"/><cs:fillRef idx="3"><cs:styleClr val="auto"/></cs:fillRef><cs:effectRef idx="3"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef></cs:dataPoint3D><cs:dataPointLine><cs:lnRef idx="0"><cs:styleClr val="auto"/></cs:lnRef><cs:fillRef idx="3"/><cs:effectRef idx="3"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:ln w="34925" cap="rnd"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:round/></a:ln></cs:spPr></cs:dataPointLine><cs:dataPointMarker><cs:lnRef idx="0"><cs:styleClr val="auto"/></cs:lnRef><cs:fillRef idx="3"><cs:styleClr val="auto"/></cs:fillRef><cs:effectRef idx="3"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:ln w="9525"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:round/></a:ln></cs:spPr></cs:dataPointMarker><cs:dataPointMarkerLayout symbol="circle" size="6"/><cs:dataPointWireframe><cs:lnRef idx="0"><cs:styleClr val="auto"/></cs:lnRef><cs:fillRef idx="3"/><cs:effectRef idx="3"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:ln w="9525" cap="rnd"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:round/></a:ln></cs:spPr></cs:dataPointWireframe><cs:dataTable><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"><a:lumMod val="85000"/></a:schemeClr></cs:fontRef><cs:spPr><a:ln w="9525"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="54000"/></a:schemeClr></a:solidFill></a:ln></cs:spPr><cs:defRPr sz="900" kern="1200"/></cs:dataTable><cs:downBar><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:solidFill><a:schemeClr val="dk1"><a:lumMod val="75000"/><a:lumOff val="25000"/></a:schemeClr></a:solidFill><a:ln w="9525"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="54000"/></a:schemeClr></a:solidFill></a:ln></cs:spPr></cs:downBar><cs:dropLine><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:ln w="9525"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="54000"/></a:schemeClr></a:solidFill><a:prstDash val="dash"/></a:ln></cs:spPr></cs:dropLine><cs:errorBar><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:ln w="9525" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/></a:schemeClr></a:solidFill><a:round/></a:ln></cs:spPr></cs:errorBar><cs:floor><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef></cs:floor><cs:gridlineMajor><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:ln w="9525" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="10000"/></a:schemeClr></a:solidFill><a:round/></a:ln></cs:spPr></cs:gridlineMajor><cs:gridlineMinor><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:ln><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="5000"/></a:schemeClr></a:solidFill></a:ln></cs:spPr></cs:gridlineMinor><cs:hiLoLine><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:ln w="9525"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="54000"/></a:schemeClr></a:solidFill><a:prstDash val="dash"/></a:ln></cs:spPr></cs:hiLoLine><cs:leaderLine><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:ln w="9525"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="54000"/></a:schemeClr></a:solidFill></a:ln></cs:spPr></cs:leaderLine><cs:legend><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"><a:lumMod val="85000"/></a:schemeClr></cs:fontRef><cs:defRPr sz="900" kern="1200"/></cs:legend><cs:plotArea><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef></cs:plotArea><cs:plotArea3D><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef></cs:plotArea3D><cs:seriesAxis><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"><a:lumMod val="85000"/></a:schemeClr></cs:fontRef><cs:spPr><a:ln w="12700" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="54000"/></a:schemeClr></a:solidFill><a:round/></a:ln></cs:spPr><cs:defRPr sz="900" kern="1200"/></cs:seriesAxis><cs:seriesLine><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:ln w="9525" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="54000"/></a:schemeClr></a:solidFill><a:round/></a:ln></cs:spPr></cs:seriesLine><cs:title><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"><a:lumMod val="95000"/></a:schemeClr></cs:fontRef><cs:defRPr sz="1600" b="1" kern="1200" spc="100" baseline="0"><a:effectLst><a:outerShdw blurRad="50800" dist="38100" dir="5400000" algn="t" rotWithShape="0"><a:prstClr val="black"><a:alpha val="40000"/></a:prstClr></a:outerShdw></a:effectLst></cs:defRPr></cs:title><cs:trendline><cs:lnRef idx="0"><cs:styleClr val="auto"/></cs:lnRef><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:ln w="19050" cap="rnd"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill></a:ln></cs:spPr></cs:trendline><cs:trendlineLabel><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"><a:lumMod val="85000"/></a:schemeClr></cs:fontRef><cs:defRPr sz="900" kern="1200"/></cs:trendlineLabel><cs:upBar><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef><cs:spPr><a:solidFill><a:schemeClr val="lt1"/></a:solidFill><a:ln w="9525"><a:solidFill><a:schemeClr val="lt1"><a:lumMod val="95000"/><a:alpha val="54000"/></a:schemeClr></a:solidFill></a:ln></cs:spPr></cs:upBar><cs:valueAxis><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"><a:lumMod val="85000"/></a:schemeClr></cs:fontRef><cs:defRPr sz="900" kern="1200"/></cs:valueAxis><cs:wall><cs:lnRef idx="0"/><cs:fillRef idx="0"/><cs:effectRef idx="0"/><cs:fontRef idx="minor"><a:schemeClr val="lt1"/></cs:fontRef></cs:wall></cs:chartStyle>'
_TPL_COLORS1       = '<cs:colorStyle xmlns:cs="http://schemas.microsoft.com/office/drawing/2012/chartStyle" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" meth="cycle" id="10"><a:schemeClr val="accent1"/><a:schemeClr val="accent2"/><a:schemeClr val="accent3"/><a:schemeClr val="accent4"/><a:schemeClr val="accent5"/><a:schemeClr val="accent6"/><cs:variation/><cs:variation><a:lumMod val="60000"/></cs:variation><cs:variation><a:lumMod val="80000"/><a:lumOff val="20000"/></cs:variation><cs:variation><a:lumMod val="80000"/></cs:variation><cs:variation><a:lumMod val="60000"/><a:lumOff val="40000"/></cs:variation><cs:variation><a:lumMod val="50000"/></cs:variation><cs:variation><a:lumMod val="70000"/><a:lumOff val="30000"/></cs:variation><cs:variation><a:lumMod val="70000"/></cs:variation><cs:variation><a:lumMod val="50000"/><a:lumOff val="50000"/></cs:variation></cs:colorStyle>'


def _applica_tema_scuro(xlsx_bytes, titoli_per_grafico):
    """
    Post-processa i grafici generati da openpyxl:
    - Sostituisce la sezione style/catAx/valAx/dTable con quella del template
    - Aggiunge style1.xml e colors1.xml per il tema scuro
    - titoli_per_grafico: dict {chart_filename: titolo}
    """
    import zipfile, re as _re

    inp = io.BytesIO(xlsx_bytes)
    out = io.BytesIO()

    with zipfile.ZipFile(inp, "r") as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        existing = set(zin.namelist())
        chart_idx = 0

        for item in zin.infolist():
            data = zin.read(item.filename)

            if item.filename.startswith("xl/charts/chart") and item.filename.endswith(".xml"):
                xml = data.decode("utf-8")

                # 1. Sostituisci AlternateContent (stile 209 scuro)
                xml = _re.sub(
                    r'<mc:AlternateContent.*?</mc:AlternateContent>',
                    _TPL_ALTCONTENT, xml, flags=_re.DOTALL
                )

                # 2. Sostituisci titolo con quello corretto per questo grafico
                chart_name = item.filename.split("/")[-1]
                titolo = titoli_per_grafico.get(chart_name, "Confronto Prezzi Carburanti")
                title_xml = _TPL_TITLE.replace("__TITLE__", titolo)
                xml = _re.sub(r'<c:title>.*?</c:title>', title_xml, xml, flags=_re.DOTALL)

                # 3. Sostituisci catAx con quello del template (etichette ruotate, stile scuro)
                xml = _re.sub(r'<c:catAx>.*?</c:catAx>', _TPL_CATAX, xml, flags=_re.DOTALL)

                # 4. Sostituisci valAx con quello del template (min=0.5, formato 0.000, stile scuro)
                xml = _re.sub(r'<c:valAx>.*?</c:valAx>', _TPL_VALAX, xml, flags=_re.DOTALL)

                # 5. Sostituisci dTable con quello del template
                xml = _re.sub(r'<c:dTable>.*?</c:dTable>', _TPL_DTABLE, xml, flags=_re.DOTALL)

                # 6. Sostituisci spPr finale del chartSpace (sfondo gradiente scuro)
                xml = _re.sub(
                    r'(</c:chart>)<c:spPr>.*?</c:spPr>',
                    r'\1' + _TPL_CHARTSPACE_SPPR, xml, flags=_re.DOTALL
                )

                data = xml.encode("utf-8")
                chart_idx += 1

                # Aggiungi style1.xml e colors1.xml per questo grafico
                chart_num = _re.search(r'chart(\d+)\.xml', item.filename)
                if chart_num:
                    n = chart_num.group(1)
                    style_path  = f"xl/charts/style{n}.xml"
                    colors_path = f"xl/charts/colors{n}.xml"
                    if style_path not in existing:
                        zout.writestr(style_path,  _TPL_STYLE1.encode())
                        existing.add(style_path)
                    if colors_path not in existing:
                        zout.writestr(colors_path, _TPL_COLORS1.encode())
                        existing.add(colors_path)

                    # Aggiorna _rels per includere style e colors
                    rels_path = f"xl/charts/_rels/chart{n}.xml.rels"
                    if rels_path in existing:
                        rels_data = zin.read(rels_path).decode()
                        if "style" not in rels_data:
                            style_rel = (
                                '<Relationship Id="rIdStyle" '
                                'Type="http://schemas.microsoft.com/office/2011/relationships/chartStyle" '
                                f'Target="style{n}.xml"/>'
                                '<Relationship Id="rIdColors" '
                                'Type="http://schemas.microsoft.com/office/2011/relationships/chartColorStyle" '
                                f'Target="colors{n}.xml"/>'
                            )
                            rels_data = rels_data.replace("</Relationships>", style_rel + "</Relationships>")
                            zout.writestr(rels_path, rels_data.encode())
                            existing.add(rels_path)
                            continue  # già scritto

            zout.writestr(item, data)

    out.seek(0)
    return out.read()


# ── Helpers per stile etichette dati ────────────────────────────────────────
def _make_label_txpr(hex_color, bold=False, italic=False):
    from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
    from openpyxl.chart.text import RichText
    rpr = CharacterProperties(b=bold or None, i=italic or None, solidFill=hex_color)
    pp  = ParagraphProperties(defRPr=rpr)
    p   = Paragraph(pPr=pp)
    return RichText(p=[p])


def _make_label_sppr(bg_hex):
    from openpyxl.chart.shapes import GraphicalProperties
    return GraphicalProperties(solidFill=bg_hex)
